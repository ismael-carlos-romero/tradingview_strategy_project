import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys

# Ensure UTF-8 output on Windows terminal
if sys.platform.startswith('win'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

# Define paths
workspace_dir = r"c:\Users\Ismael Romero\.gemini\antigravity\scratch\tradingview_strategy_project"
downloads_dir = r"C:\Users\Ismael Romero\Downloads"
json_filename = "ucharts_extracted_trades.json"
excel_filename = "PROYECTO MELIORA OPCIONES FINANCIERAS.xlsx"

# Find JSON file
json_path = os.path.join(workspace_dir, json_filename)
if not os.path.exists(json_path):
    json_path = os.path.join(downloads_dir, json_filename)

if not os.path.exists(json_path):
    print(f"Error: No se encontro '{json_filename}' ni en el proyecto ni en Descargas.")
    print("Por favor ejecuta primero el script de consola en UCharts para descargar el archivo.")
    exit(1)

print(f"Cargando datos de: {json_path}")
with open(json_path, 'r', encoding='utf-8') as f:
    trades = json.load(f)

print(f"Se cargaron {len(trades)} operaciones del archivo JSON.")

# Open excel workbook
excel_path = os.path.join(workspace_dir, excel_filename)
if not os.path.exists(excel_path):
    print(f"Error: No se encontro '{excel_filename}' en el proyecto.")
    exit(1)

print(f"Abriendo archivo Excel: {excel_path}")
# Make a backup first
backup_path = excel_path.replace(".xlsx", "_backup.xlsx")
import shutil
shutil.copyfile(excel_path, backup_path)
print(f"Se creo una copia de seguridad en: {backup_path}")

wb = openpyxl.load_workbook(excel_path)

# Create or clear "UCharts Compuesto" sheet
sheet_name = "UCharts Compuesto"
if sheet_name in wb.sheetnames:
    print(f"Limpiando pestaña existente: {sheet_name}")
    ws = wb[sheet_name]
    wb.remove(ws)
ws = wb.create_sheet(title=sheet_name)

# Set grid lines visible
ws.views.sheetView[0].showGridLines = True

# Add title block (mirroring user's style)
ws['A1'] = "UCHARTS - CAPITAL COMPUESTO"
ws['A1'].font = Font(name="Arial", size=14, bold=True)
ws.merge_cells("A1:G1")

# Create summary labels as placeholders (user can copy formulas from other sheets)
summary_data = [
    ("Capital inicial", 1000.00, "Capital actual", 1000.00, "Trades cerrados", 0.00),
    ("% max por trade", 0.10, "Cuenta personal", 0.00, "Ganadores", 0.00),
    ("% retiro sobre ganancia", 0.10, "Ganancia/Perdida acumulada", 0.00, "Perdedores", 0.00),
    ("Comision por trade", 2.10, "Retirado total", 0.00, "Resultado del mes", 0.00),
    ("Mes a ver", "jun 2026", "Regla", "Se reinvierte el capital disponible; solo se retira % de la ganancia cuando el resultado es positivo.")
]

for idx, row in enumerate(summary_data):
    r = idx + 2
    ws.cell(row=r, column=1, value=row[0]).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=r, column=2, value=row[1])
    ws.cell(row=r, column=4, value=row[2]).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=r, column=5, value=row[3])
    if len(row) > 4:
        ws.cell(row=r, column=6, value=row[4]).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=r, column=7, value=row[5])

# Add header row at row 8
headers = [
    "ID", "Estado", "Fecha compra", "Fecha venta/cierre", "Mes", "Semana", "Dia", 
    "Ticker", "Tipo", "Strike", "Vencimiento", "Capital antes", "% máx inversión", 
    "Inversión sugerida", "Cantidad", "Prima compra", "Costo real compra", 
    "TP % pretendido", "Precio TP pretendido", "Prima venta", "Estrategia"
]

header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=8, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

# Mappings (1-based index)
col_map = {h.lower(): idx for idx, h in enumerate(headers, 1)}

# Helper to convert index to excel column letter
def get_col_letter(col_idx):
    return openpyxl.utils.get_column_letter(col_idx)

print("Escribiendo transacciones a la hoja...")

# Format styles for data rows
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

for i, trade in enumerate(trades):
    row_idx = 9 + i
    
    # 1. ID
    ws.cell(row=row_idx, column=col_map["id"], value=i+1).alignment = Alignment(horizontal="center")
    
    # 2. Estado
    ws.cell(row=row_idx, column=col_map["estado"], value=trade.get("status", "Cerrado")).alignment = Alignment(horizontal="center")
    
    # 3. Fecha compra
    ws.cell(row=row_idx, column=col_map["fecha compra"], value=trade.get("buyDate", "")).alignment = Alignment(horizontal="center")
    
    # 4. Fecha venta/cierre
    ws.cell(row=row_idx, column=col_map["fecha venta/cierre"], value=trade.get("sellDate", "")).alignment = Alignment(horizontal="center")
    
    # 5. Ticker
    ws.cell(row=row_idx, column=col_map["ticker"], value=trade.get("ticker", "")).alignment = Alignment(horizontal="center")
    
    # 6. Tipo
    ws.cell(row=row_idx, column=col_map["tipo"], value=trade.get("type", "")).alignment = Alignment(horizontal="center")
    
    # 7. Strike
    ws.cell(row=row_idx, column=col_map["strike"], value=trade.get("strike", ""))
    
    # 8. Vencimiento
    ws.cell(row=row_idx, column=col_map["vencimiento"], value=trade.get("expiry", "")).alignment = Alignment(horizontal="center")
    
    # 9. Cantidad
    ws.cell(row=row_idx, column=col_map["cantidad"], value=trade.get("quantity", 1))
    
    # 10. Prima compra
    ws.cell(row=row_idx, column=col_map["prima compra"], value=trade.get("buyPrice", 0.0))
    
    # 11. Prima venta
    ws.cell(row=row_idx, column=col_map["prima venta"], value=trade.get("sellPrice", 0.0))
    
    # 12. Estrategia
    ws.cell(row=row_idx, column=col_map["estrategia"], value=trade.get("strategy", ""))
    
    # 13. Formulas
    f_buy_date_col = get_col_letter(col_map["fecha compra"])
    f_qty_col = get_col_letter(col_map["cantidad"])
    f_buy_price_col = get_col_letter(col_map["prima compra"])
    
    # Mes
    ws.cell(row=row_idx, column=col_map["mes"], value=f'=IF({f_buy_date_col}{row_idx}<>"", TEXT({f_buy_date_col}{row_idx}, "mmmm"), "")')
    
    # Semana
    ws.cell(row=row_idx, column=col_map["semana"], value=f'=IF({f_buy_date_col}{row_idx}<>"", WEEKNUM({f_buy_date_col}{row_idx}, 2), "")')
    
    # Dia
    ws.cell(row=row_idx, column=col_map["dia"], value=f'=IF({f_buy_date_col}{row_idx}<>"", TEXT({f_buy_date_col}{row_idx}, "dddd"), "")')
    
    # Costo Real
    ws.cell(row=row_idx, column=col_map["costo real compra"], value=f'=IF(AND({f_qty_col}{row_idx}<>"", {f_buy_price_col}{row_idx}<>Normalized), {f_qty_col}{row_idx}*{f_buy_price_col}{row_idx}*100, "")'.replace("Normalized", '""'))
    
    # Borders
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=row_idx, column=col_idx).border = thin_border

# Autofit columns
for col in ws.columns:
    max_len = 0
    col_letter = openpyxl.utils.get_column_letter(col[0].column)
    for cell in col:
        if cell.row < 8:
            continue
        val = str(cell.value or '')
        if val.startswith('='):
            val = "Formula"
        max_len = max(max_len, len(val))
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

print("Guardando archivo...")
wb.save(excel_path)
print(f"Exito! Se guardaron {len(trades)} operaciones en '{excel_path}' bajo la pestaña '{sheet_name}'.")
