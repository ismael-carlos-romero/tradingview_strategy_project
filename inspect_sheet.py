import openpyxl
import pandas as pd

wb = openpyxl.load_workbook('PROYECTO MELIORA OPCIONES FINANCIERAS.xlsx', read_only=True)
with open('inspect_results.txt', 'w', encoding='utf-8') as out:
    out.write("Sheet Names: " + str(wb.sheetnames) + "\n")
    
    # Check sheet names with 'uchart'
    uchart_sheets = [s for s in wb.sheetnames if 'uchart' in s.lower() or 'compuesto' in s.lower()]
    out.write("Matching sheets: " + str(uchart_sheets) + "\n")
    
    for sheet in uchart_sheets:
        out.write(f"\n--- Columns in {sheet} ---\n")
        df = pd.read_excel('PROYECTO MELIORA OPCIONES FINANCIERAS.xlsx', sheet_name=sheet)
        out.write("Shape: " + str(df.shape) + "\n")
        out.write("Columns: " + str(list(df.columns)) + "\n")
        out.write("Head:\n" + df.head(15).to_string() + "\n")

